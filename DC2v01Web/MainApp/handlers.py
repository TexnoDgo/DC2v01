import os
from openpyxl import Workbook
import openpyxl
from random import random

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# Создание производственной таблицы для аутсорсеров
def create_operation_table(operation_lists, send_order):
    try:
        print('start to create table')
        # Создание таблицы
        wb = Workbook()
        ws = wb.active

        # Поля проекта. Шапка
        ws['A1'] = 'ЗАКАЗ'
        ws['B1'] = 'ПРОЕКТ'

        # Поля проекта. Заполнение
        ws['A2'] = send_order.title
        ws['B2'] = send_order.project.title

        i = 4

        ws['A3'] = "#"
        ws['B3'] = "Имя компонента"
        ws['C3'] = "Тип компонента"
        ws['D3'] = "Материал"
        ws['E3'] = "Толщина, мм"
        ws['F3'] = "Кол-во гибов"
        ws['G3'] = "Кол-во деталей, шт"
        ws['H3'] = "Примечание"

        for one_operation in operation_lists:

            ws.cell(row=i, column=1, value=one_operation.position.pk)

            ws.cell(row=i, column=2, value=one_operation.position.component.title)

            ws.cell(row=i, column=3, value=one_operation.position.component.c_type)

            ws.cell(row=i, column=4, value=one_operation.position.component.material)

            ws.cell(row=i, column=5, value=one_operation.position.component.thickness)

            ws.cell(row=i, column=6, value=one_operation.position.component.band_count)

            ws.cell(row=i, column=7, value=one_operation.position.quantity)

            ws.cell(row=i, column=8, value=one_operation.notes)


            i += 1

        a = random()
        b = a * 10000000000000
        c = str(b)[:13]

        table_name = 'Таблица заказка - #' + str(c)

        table_path = os.path.join(BASE_DIR, 'media/temp/') + "{}.xlsx".format(table_name)

        wb.save(table_path)

        return table_path.replace('/', '\\')

    except Exception:
        pass
