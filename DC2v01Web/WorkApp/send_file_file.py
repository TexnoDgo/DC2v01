import os
from wsgiref.util import FileWrapper
from openpyxl import Workbook
import openpyxl

from django.http import HttpResponse
from django.conf import settings

from .models import Order, Project, Position, Component

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# Функции отправки файла компонента
def send_comp_draw_pdf(component_name):
    try:
        component = Component.objects.get(title=component_name)
        draw = component.draw_pdf
        response = HttpResponse(FileWrapper(draw), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % '111.pdf'
        return response
    except Exception:
        print("Ошибка отправки PDF чертежа компонента")
        return False


def send_comp_draw_png(component_name):
    try:
        print("Начат процесс получение png чертежа компонента")
        component = Component.objects.get(title=component_name)
        draw = component.draw_pdf
        response = HttpResponse(FileWrapper(draw), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % '111.pdf'
        print("Процесс завершен!")
        return response
    except Exception:
        print("Ошибка отправки PNG чертежа компонента")
        return False


# Формирование файла спецификации сборки
def assembly_spec_forming(project_name, assembly_name):
    try:
        print("1")
        project = Project.objects.get(title=project_name)
        orders = Order.objects.filter(project=project)
        # created xls document
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        # ORDER CAP
        #
        # Поля проекта. Шапка
        ws['A1'] = 'ПРОЕКТ'
        ws['B1'] = 'НАЗВАНИЕ'
        ws['C1'] = 'АВТОР'
        ws['D1'] = 'СОЗДАН'
        ws['E1'] = 'УСТРОЙСТВО'
        # Поля проекта. Заполнение
        ws['A2'] = project.pk
        ws['B2'] = project.title
        ws['C2'] = project.author.username
        ws['D2'] = project.create
        ws['E2'] = project.device.title
        ws.merge_cells('A3:E3')
        ws['A3'] = assembly_name
        i = 4
        print("2")
        for order in orders:
            m_c = "A{0}:L{0}".format(i)
            ws.merge_cells(m_c)
            m_c_s = "A{0}".format(i)
            ws[m_c_s] = order.title

            # Поля проекта. Шапка
            '''ws.cell(row=i, column=1, value="#")
            ws.cell(row=i, column=2, value="Имя компонента")
            ws.cell(row=i, column=3, value="Тип компонента")
            ws.cell(row=i, column=4, value="Материал")
            ws.cell(row=i, column=5, value="Толщина, мм")
            ws.cell(row=i, column=6, value="Кол-во гибов")
            ws.cell(row=i, column=7, value="Кол-во деталей, шт")
            ws.cell(row=i, column=8, value="Приоритет")
            ws.cell(row=i, column=9, value="Место хранения")
            ws.cell(row=i, column=10, value="QR-код ")
            ws.cell(row=i, column=11, value="QR-код производства")
            ws.cell(row=i, column=12, value="Чертеж")'''
            a_0 = "A{}".format(i+1)
            b_0 = "B{}".format(i+1)
            c_0 = "C{}".format(i+1)
            d_0 = "D{}".format(i+1)
            e_0 = "E{}".format(i+1)
            f_0 = "F{}".format(i+1)
            g_0 = "G{}".format(i+1)
            h_0 = "H{}".format(i+1)
            i_0 = "I{}".format(i+1)
            j_0 = "J{}".format(i+1)
            k_0 = "K{}".format(i+1)
            l_0 = "L{}".format(i+1)
            ws[a_0] = "#"
            ws[b_0] = "Имя компонента"
            ws[c_0] = "Тип компонента"
            ws[d_0] = "Материал"
            ws[e_0] = "Толщина, мм"
            ws[f_0] = "Кол-во гибов"
            ws[g_0] = "Кол-во деталей, шт"
            ws[h_0] = "Приоритет"
            ws[i_0] = "Место хранения"
            ws[j_0] = "QR-код"
            ws[k_0] = "QR-код производства"
            ws[l_0] = "Чертеж"

            positions = Position.objects.filter(order=order, mather_assembly=assembly_name)
            for position in positions:
                print("-----------------------")
                print(position.component.title)
                print(position.quantity)
                print(position.priority)
                print(position.storage)
                print("-----------------------")
                ws.cell(row=i+2, column=1, value=position.pk)
                ws.cell(row=i+2, column=2, value=position.component.title)
                ws.cell(row=i+2, column=3, value=position.component.c_type)
                ws.cell(row=i+2, column=4, value=position.component.material)
                ws.cell(row=i+2, column=5, value=position.component.thickness)
                ws.cell(row=i+2, column=6, value=position.component.band_count)
                ws.cell(row=i+2, column=7, value=position.quantity)
                ws.cell(row=i+2, column=8, value=position.priority)
                ws.cell(row=i+2, column=9, value=position.storage)
                qr_0 = "J{}".format(i+2)
                qr_p_0 = "K{}".format(i+2)
                dr_0 = "L{}".format(i+2)

                k = 100
                y = 200

                qr_code = openpyxl.drawing.image.Image(position.qr_code.path)
                qr_code.width = k
                qr_code.height = k
                qr_code_production = openpyxl.drawing.image.Image(position.qr_code_production.path)
                qr_code_production.width = k
                qr_code_production.height = k
                draw = openpyxl.drawing.image.Image(position.component.draw_png.path)
                draw.width = y
                draw.height = y

                '''ws.cell(row=i+2, column=10, value=qr_code)
                ws.cell(row=i+2, column=11, value=qr_code_production)
                ws.cell(row=i+2, column=12, value=draw)'''

                '''qr_code.anchor(ws.cell(row=i+2, column=7))
                qr_code_production.anchor(ws.cell(row=i+2, column=8))
                draw.anchor(ws.cell(row=i+2, column=9))'''
                print("3")
                '''ws.add_image(qr_code)
                ws.add_image(qr_code_production)
                ws.add_image(draw)'''
                ws.add_image(qr_code, qr_0)
                ws.add_image(qr_code_production, qr_p_0)
                ws.add_image(draw, dr_0)

                i += 1

        table_path = os.path.join(BASE_DIR, 'media/temp/') + "{}.xlsx".format(project.title)
        obj = wb.save(table_path)
        with open(table_path, 'rb') as ft:
            response = HttpResponse(ft.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(table_path)
            return response
    except Exception:
        return False